#!/usr/bin/env python3
"""Migrate the pre-1.0.14 API case project into an Arazzo workflow document."""
from __future__ import annotations

import csv
import json
import os
import re
import tempfile
import urllib.parse
from collections import OrderedDict
from pathlib import Path
from typing import Any

import yaml


class LegacyMigrationError(ValueError):
    """Raised when legacy semantics cannot be preserved."""


_TEMPLATE = re.compile(r"\$\{(ENV|vars|project|steps)\.([^}]+)\}")
_STEP_ID = re.compile(r"[^A-Za-z0-9_-]+")
_METHODS = {"get", "put", "post", "delete", "options", "head", "patch", "trace"}
_TABLE_SUFFIXES = {".csv", ".xlsx", ".xls"}
_MAP_SUFFIXES = {".yaml", ".yml", ".json"}


def _load_mapping(path: Path) -> dict[str, Any]:
    try:
        value = (
            json.loads(path.read_text(encoding="utf-8"))
            if path.suffix.lower() == ".json"
            else yaml.safe_load(path.read_text(encoding="utf-8"))
        )
    except (OSError, UnicodeError, json.JSONDecodeError, yaml.YAMLError) as exc:
        raise LegacyMigrationError(f"Unable to read {path}: {exc}") from exc
    if not isinstance(value, dict):
        raise LegacyMigrationError(f"Expected an object in {path}")
    return value


def _step_id(value: object, fallback: str) -> str:
    result = _STEP_ID.sub("_", str(value or fallback)).strip("_")
    if not result:
        result = fallback
    return result


def _operation_index(schema: dict[str, Any]) -> list[tuple[str, str, str]]:
    result: list[tuple[str, str, str]] = []
    for path, path_item in (schema.get("paths") or {}).items():
        if not isinstance(path_item, dict):
            continue
        for method, operation in path_item.items():
            if method.lower() not in _METHODS or not isinstance(operation, dict):
                continue
            operation_id = operation.get("operationId")
            if isinstance(operation_id, str) and operation_id:
                result.append((method.upper(), str(path), operation_id))
    return result


def _segments_match(legacy_path: str, schema_path: str) -> bool:
    left = legacy_path.strip("/").split("/")
    right = schema_path.strip("/").split("/")
    if len(left) != len(right):
        return False
    for legacy, schema in zip(left, right):
        if schema.startswith("{") and schema.endswith("}"):
            continue
        if legacy != schema:
            return False
    return True


def _find_operation(
    method: str,
    raw_url: str,
    operations: list[tuple[str, str, str]],
) -> tuple[str, str]:
    parsed = urllib.parse.urlsplit(raw_url)
    if parsed.scheme or parsed.netloc:
        raise LegacyMigrationError(
            f"Legacy request URL must be relative to the project base URL: {raw_url}"
        )
    candidates = [
        (schema_path, operation_id)
        for candidate_method, schema_path, operation_id in operations
        if candidate_method == method.upper() and _segments_match(parsed.path, schema_path)
    ]
    if len(candidates) != 1:
        detail = "no" if not candidates else "multiple"
        raise LegacyMigrationError(
            f"{detail} OpenAPI operation matches {method.upper()} {raw_url}"
        )
    return candidates[0]


def _translate_expression(
    value: Any,
    *,
    required_env: set[str],
    project_vars: dict[str, Any],
) -> Any:
    if isinstance(value, list):
        return [
            _translate_expression(
                item, required_env=required_env, project_vars=project_vars
            )
            for item in value
        ]
    if isinstance(value, dict):
        return {
            key: _translate_expression(
                item, required_env=required_env, project_vars=project_vars
            )
            for key, item in value.items()
        }
    if not isinstance(value, str):
        return value

    matches = list(_TEMPLATE.finditer(value))
    if not matches:
        return value

    def converted(kind: str, path: str) -> str:
        if kind == "ENV":
            required_env.add(path)
            return f"$inputs.{path}"
        if kind == "vars":
            return f"$outputs.{path}"
        if kind == "project":
            if path not in project_vars:
                raise LegacyMigrationError(f"Unknown project variable: {path}")
            return f"$inputs.project_{path}"
        step_name, separator, output_path = path.partition(".")
        if not separator:
            raise LegacyMigrationError(f"Incomplete legacy step expression: {path}")
        if output_path == "status_code":
            return f"$steps.{_step_id(step_name, 'step')}.outputs.status_code"
        if output_path == "json":
            return f"$steps.{_step_id(step_name, 'step')}.outputs.body"
        if output_path.startswith("json."):
            pointer = "/" + "/".join(output_path[5:].split("."))
            return (
                f"$steps.{_step_id(step_name, 'step')}.outputs.body#{pointer}"
            )
        raise LegacyMigrationError(f"Unsupported legacy step expression: {path}")

    if len(matches) == 1 and matches[0].span() == (0, len(value)):
        return converted(matches[0].group(1), matches[0].group(2))
    pieces: list[str] = []
    position = 0
    for match in matches:
        pieces.append(value[position : match.start()])
        pieces.append("{" + converted(match.group(1), match.group(2)) + "}")
        position = match.end()
    pieces.append(value[position:])
    return "".join(pieces)


def _field_expression(field: Any) -> str:
    if field == "status_code":
        return "$statusCode"
    if isinstance(field, str) and field.startswith("$."):
        return "$response.body#/" + "/".join(field[2:].split("."))
    if isinstance(field, str) and field in {"text", "headers"}:
        raise LegacyMigrationError(
            f"Legacy assertion field {field!r} has no lossless Arazzo mapping"
        )
    raise LegacyMigrationError(f"Unsupported legacy assertion field: {field!r}")


def _criteria(
    validations: Any,
    *,
    required_env: set[str],
    project_vars: dict[str, Any],
) -> list[dict[str, Any]]:
    if validations is None:
        return []
    if not isinstance(validations, list):
        raise LegacyMigrationError("Legacy validate must be a list")
    result: list[dict[str, Any]] = []
    operators = {
        "eq": "==",
        "ne": "!=",
        "gt": ">",
        "gte": ">=",
        "lt": "<",
        "lte": "<=",
    }
    for validation in validations:
        if not isinstance(validation, dict) or len(validation) != 1:
            raise LegacyMigrationError(f"Invalid legacy validation: {validation!r}")
        name, operands = next(iter(validation.items()))
        if name in operators:
            if not isinstance(operands, list) or len(operands) != 2:
                raise LegacyMigrationError(f"Invalid {name} validation")
            field, expected = operands
            expected = _translate_expression(
                expected, required_env=required_env, project_vars=project_vars
            )
            rendered = (
                expected if isinstance(expected, str) and expected.startswith("$")
                else json.dumps(expected, ensure_ascii=False)
            )
            result.append(
                {
                    "condition": (
                        f"{_field_expression(field)} {operators[name]} {rendered}"
                    )
                }
            )
        elif name == "contains":
            if not isinstance(operands, list) or len(operands) != 2:
                raise LegacyMigrationError("Invalid contains validation")
            field, expected = operands
            expected = _translate_expression(
                expected, required_env=required_env, project_vars=project_vars
            )
            result.append(
                {
                    "type": "x-testkit-contains",
                    "context": _field_expression(field),
                    "condition": (
                        expected
                        if isinstance(expected, str) and expected.startswith("$")
                        else json.dumps(expected, ensure_ascii=False)
                    ),
                }
            )
        elif name == "exists":
            result.append(
                {
                    "type": "x-testkit-exists",
                    "condition": _field_expression(operands),
                }
            )
        else:
            raise LegacyMigrationError(f"Unsupported legacy validation: {name}")
    return result


def _path_parameters(
    legacy_url: str,
    schema_path: str,
    *,
    required_env: set[str],
    project_vars: dict[str, Any],
) -> list[dict[str, Any]]:
    legacy_segments = urllib.parse.urlsplit(legacy_url).path.strip("/").split("/")
    schema_segments = schema_path.strip("/").split("/")
    result: list[dict[str, Any]] = []
    for legacy, schema in zip(legacy_segments, schema_segments):
        if schema.startswith("{") and schema.endswith("}"):
            name = schema[1:-1]
            result.append(
                {
                    "name": name,
                    "in": "path",
                    "value": _translate_expression(
                        urllib.parse.unquote(legacy),
                        required_env=required_env,
                        project_vars=project_vars,
                    ),
                }
            )
    return result


def _convert_step(
    raw: Any,
    *,
    index: int,
    operations: list[tuple[str, str, str]],
    required_env: set[str],
    project_vars: dict[str, Any],
    default_headers: dict[str, Any],
) -> dict[str, Any]:
    if not isinstance(raw, dict):
        raise LegacyMigrationError(f"Legacy step {index} must be an object")
    step_id = _step_id(raw.get("name") or raw.get("save_as"), f"step_{index + 1}")
    delay_ms: int | None = None
    if raw.get("sleep") is not None:
        sleep_value = raw["sleep"]
        if (
            not isinstance(sleep_value, (int, float))
            or isinstance(sleep_value, bool)
            or sleep_value < 0
            or sleep_value > 300
        ):
            raise LegacyMigrationError(
                f"Step {step_id} sleep must be between 0 and 300 seconds"
            )
        delay_ms = int(sleep_value * 1000)
    if delay_ms is not None and not raw.get("request") and not raw.get("use"):
        return {
            "stepId": step_id,
            "x-testkit-delay-ms": delay_ms,
        }
    if raw.get("use"):
        use = raw["use"]
        if not isinstance(use, str) or not use.startswith("flow:"):
            raise LegacyMigrationError(f"Unsupported flow reference: {use!r}")
        parameters = [
            {
                "name": str(name),
                "in": "workflow",
                "value": _translate_expression(
                    value,
                    required_env=required_env,
                    project_vars=project_vars,
                ),
            }
            for name, value in (raw.get("inputs") or {}).items()
        ]
        return {
            "stepId": step_id,
            "workflowId": use.split(":", 1)[1],
            **(
                {"x-testkit-delay-ms": delay_ms}
                if delay_ms is not None
                else {}
            ),
            **({"parameters": parameters} if parameters else {}),
        }
    request = raw.get("request")
    if not isinstance(request, dict):
        raise LegacyMigrationError(f"Step {step_id} requires request or use")
    method = str(request.get("method", "GET")).upper()
    raw_url = request.get("url")
    if not isinstance(raw_url, str) or not raw_url:
        raise LegacyMigrationError(f"Step {step_id} requires request.url")
    schema_path, operation_id = _find_operation(method, raw_url, operations)
    parameters = _path_parameters(
        raw_url,
        schema_path,
        required_env=required_env,
        project_vars=project_vars,
    )
    parsed = urllib.parse.urlsplit(raw_url)
    for name, value in urllib.parse.parse_qsl(parsed.query, keep_blank_values=True):
        parameters.append(
            {
                "name": name,
                "in": "query",
                "value": _translate_expression(
                    value, required_env=required_env, project_vars=project_vars
                ),
            }
        )
    for name, value in (request.get("params") or {}).items():
        parameters.append(
            {
                "name": str(name),
                "in": "query",
                "value": _translate_expression(
                    value, required_env=required_env, project_vars=project_vars
                ),
            }
        )
    headers = {**default_headers, **(request.get("headers") or {})}
    for name, value in headers.items():
        parameters.append(
            {
                "name": str(name),
                "in": "header",
                "value": _translate_expression(
                    value, required_env=required_env, project_vars=project_vars
                ),
            }
        )
    request_body: dict[str, Any] | None = None
    if "json" in request:
        request_body = {
            "contentType": "application/json",
            "payload": _translate_expression(
                request["json"],
                required_env=required_env,
                project_vars=project_vars,
            ),
        }
    elif "data" in request or "raw_body" in request:
        payload = request.get("data", request.get("raw_body"))
        request_body = {
            "contentType": (
                "application/x-www-form-urlencoded"
                if isinstance(payload, dict)
                else "text/plain"
            ),
            "payload": _translate_expression(
                payload, required_env=required_env, project_vars=project_vars
            ),
        }
    outputs: dict[str, Any] = {
        "status_code": "$statusCode",
        "body": "$response.body",
    }
    extracts = raw.get("extract") or {}
    if not isinstance(extracts, dict):
        raise LegacyMigrationError(f"Step {step_id} extract must be an object")
    outputs.update(
        {
            str(name): {
                "type": "jsonpath",
                "context": "$response.body",
                "selector": selector,
            }
            for name, selector in extracts.items()
        }
    )
    timeout = request.get("timeout", 30)
    if not isinstance(timeout, (int, float)) or isinstance(timeout, bool):
        raise LegacyMigrationError(f"Step {step_id} timeout must be numeric")
    return {
        "stepId": step_id,
        "operationId": operation_id,
        **({"parameters": parameters} if parameters else {}),
        **({"requestBody": request_body} if request_body else {}),
        "successCriteria": _criteria(
            raw.get("validate"),
            required_env=required_env,
            project_vars=project_vars,
        ),
        "outputs": outputs,
        "timeout": int(timeout * 1000),
        **(
            {"x-testkit-delay-ms": delay_ms}
            if delay_ms is not None
            else {}
        ),
        **(
            {"x-testkit-aliases": [str(raw["save_as"])]}
            if raw.get("save_as") and raw["save_as"] != raw.get("name")
            else {}
        ),
    }


def _tabular_cases(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
    else:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise LegacyMigrationError("openpyxl is required for Excel migration") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if not values:
            return []
        headers = [str(item or "").strip() for item in values[0]]
        rows = [
            {
                headers[index]: "" if item is None else str(item)
                for index, item in enumerate(row)
                if index < len(headers) and headers[index]
            }
            for row in values[1:]
        ]
    groups: OrderedDict[str, list[dict[str, str]]] = OrderedDict()
    for row_index, row in enumerate(rows, start=2):
        if not any(str(value or "").strip() for value in row.values()):
            continue
        case_id = str(row.get("用例ID") or "").strip()
        if not case_id:
            raise LegacyMigrationError(
                f"Tabular case row {row_index} is missing 用例ID"
            )
        groups.setdefault(case_id, []).append(row)
    cases: list[dict[str, Any]] = []
    for case_id, group in groups.items():
        first_name = str(group[0].get("用例名称") or case_id)
        case = {
            "id": case_id,
            "name": first_name.split(" / ", 1)[0],
            "setup": [],
            "steps": [],
            "teardown": [],
        }
        for row in group:
            flow = str(row.get("前置依赖") or "").strip()
            url = str(row.get("接口路径") or "").strip()
            name = str(row.get("用例名称") or "step").split(" / ", 1)[-1]
            if flow and not url:
                step: dict[str, Any] = {"name": name, "use": flow}
            else:
                body = str(row.get("请求体/参数") or "").strip()
                request: dict[str, Any] = {
                    "method": str(row.get("请求方法") or "GET").strip(),
                    "url": url,
                }
                if body:
                    try:
                        request["json"] = json.loads(body)
                    except json.JSONDecodeError:
                        request["raw_body"] = body
                validations: list[dict[str, Any]] = []
                status = str(row.get("预期状态码") or "").strip()
                if status:
                    validations.append({"eq": ["status_code", int(status)]})
                checks = str(row.get("预期响应校验") or "").strip()
                if checks:
                    validations.extend(_parse_table_checks(checks))
                step = {"name": name, "request": request, "validate": validations}
                extracts = _parse_table_extracts(
                    str(row.get("依赖产出提取") or "")
                )
                if extracts:
                    step["extract"] = extracts
            phase = str(row.get("注入方式") or "").strip().lower()
            case[phase if phase in {"setup", "teardown"} else "steps"].append(step)
        cases.append(case)
    return cases


def _parse_table_checks(value: str) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    pattern = re.compile(r"(\$[\w.]+)\s*(==|!=|>=|<=|>|<|\s+contains\s+)\s*(.+)")
    names = {"==": "eq", "!=": "ne", ">": "gt", ">=": "gte", "<": "lt", "<=": "lte"}
    for part in (item.strip() for item in value.split("&&")):
        match = pattern.fullmatch(part)
        if not match:
            result.append({"exists": part})
            continue
        field, operator, expected = match.groups()
        operator = operator.strip()
        try:
            typed = json.loads(expected)
        except json.JSONDecodeError:
            typed = expected
        result.append(
            {
                "contains" if operator == "contains" else names[operator]: [
                    field,
                    typed,
                ]
            }
        )
    return result


def _parse_table_extracts(value: str) -> dict[str, str]:
    result: dict[str, str] = {}
    for part in value.split(";"):
        name, separator, selector = part.partition("=")
        if separator and name.strip() and selector.strip():
            result[name.strip()] = selector.strip()
    return result


def _load_legacy_documents(
    flows_root: Path,
    cases_root: Path,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    flows: dict[str, Any] = {}
    cases: list[dict[str, Any]] = []
    for path in sorted(flows_root.rglob("*")) if flows_root.is_dir() else []:
        if path.suffix.lower() in _MAP_SUFFIXES:
            raw = _load_mapping(path)
            current = raw.get("flows") or {}
            if not isinstance(current, dict):
                raise LegacyMigrationError(f"flows must be an object in {path}")
            duplicate = set(flows) & set(current)
            if duplicate:
                raise LegacyMigrationError(
                    f"Duplicate flow names: {', '.join(sorted(duplicate))}"
                )
            flows.update(current)
    for path in sorted(cases_root.rglob("*")) if cases_root.is_dir() else []:
        if path.suffix.lower() in _MAP_SUFFIXES:
            document = _load_mapping(path)
            inline_flows = document.get("flows") or {}
            if not isinstance(inline_flows, dict):
                raise LegacyMigrationError(f"flows must be an object in {path}")
            duplicate = set(flows) & set(inline_flows)
            if duplicate:
                raise LegacyMigrationError(
                    f"Duplicate flow names: {', '.join(sorted(duplicate))}"
                )
            flows.update(inline_flows)
            raw_cases = document.get("cases") or []
        elif path.suffix.lower() in _TABLE_SUFFIXES:
            raw_cases = _tabular_cases(path)
        else:
            continue
        if not isinstance(raw_cases, list):
            raise LegacyMigrationError(f"cases must be a list in {path}")
        cases.extend(raw_cases)
    return flows, cases


def _resolve_project_directory(project_root: Path, value: Any, *, field: str) -> Path:
    project_root = project_root.resolve()
    candidate = (project_root / str(value)).resolve()
    try:
        candidate.relative_to(project_root)
    except ValueError as exc:
        raise LegacyMigrationError(
            f"project.report.{field} must stay inside the legacy project directory"
        ) from exc
    return candidate


def migrate_legacy_project(
    project_path: Path,
    schema_path: Path,
    output_path: Path,
) -> dict[str, Any]:
    """Convert a complete legacy project atomically and return a manifest."""
    project_path = project_path.resolve()
    schema_path = schema_path.resolve()
    output_path = output_path.resolve()
    project_doc = _load_mapping(project_path)
    project = project_doc.get("project", project_doc)
    if not isinstance(project, dict):
        raise LegacyMigrationError("project must be an object")
    project_vars = project.get("vars") or {}
    default_headers = (project.get("defaults") or {}).get("headers") or {}
    if not isinstance(project_vars, dict) or not isinstance(default_headers, dict):
        raise LegacyMigrationError("project vars and default headers must be objects")
    operations = _operation_index(_load_mapping(schema_path))
    report = project.get("report") or {}
    if not isinstance(report, dict):
        raise LegacyMigrationError("project report must be an object")
    project_root = project_path.parent
    flows_root = _resolve_project_directory(
        project_root,
        report.get("flows_dir", "flows"),
        field="flows_dir",
    )
    cases_root = _resolve_project_directory(
        project_root,
        report.get("cases_dir", "cases"),
        field="cases_dir",
    )
    flows, cases = _load_legacy_documents(flows_root, cases_root)
    workflow_rows: list[tuple[dict[str, Any], set[str]]] = []

    def convert_workflow(
        workflow_id: str,
        raw: dict[str, Any],
        *,
        is_case: bool,
    ) -> None:
        local_required_env: set[str] = set()
        phases: dict[str, list[dict[str, Any]]] = {}
        for legacy_name, target_name in (
            ("setup", "x-testkit-setup"),
            ("steps", "steps"),
            ("teardown", "x-testkit-cleanup"),
        ):
            converted = [
                _convert_step(
                    step,
                    index=index,
                    operations=operations,
                    required_env=local_required_env,
                    project_vars=project_vars,
                    default_headers=default_headers,
                )
                for index, step in enumerate(raw.get(legacy_name, []))
            ]
            if converted or target_name == "steps":
                phases[target_name] = converted
        extracted = sorted(
            {
                name
                for phase in phases.values()
                for step in phase
                for name in (step.get("outputs") or {})
                if name not in {"status_code", "body"}
            }
        )
        workflow = {
            "workflowId": workflow_id,
            "summary": str(raw.get("name") or workflow_id),
            **(
                {"x-testkit-tags": list(raw.get("tags") or [])}
                if is_case and raw.get("tags")
                else {}
            ),
            **(
                {"x-testkit-continue-on-failure": True}
                if raw.get("continue_on_failure")
                or raw.get("on_failure") == "continue"
                else {}
            ),
            **phases,
            **(
                {
                    "outputs": {
                        name: f"$outputs.{name}"
                        for name in extracted
                    }
                }
                if extracted
                else {}
            ),
        }
        workflow_rows.append((workflow, local_required_env))

    for name, raw in flows.items():
        if not isinstance(raw, dict):
            raise LegacyMigrationError(f"Flow {name} must be an object")
        convert_workflow(name, raw, is_case=False)
    seen_ids: set[str] = set(flows)
    for index, raw in enumerate(cases):
        if not isinstance(raw, dict):
            raise LegacyMigrationError(f"Case {index} must be an object")
        case_id = _step_id(raw.get("id"), f"case_{index + 1}")
        if case_id in seen_ids:
            raise LegacyMigrationError(f"Duplicate workflow id: {case_id}")
        seen_ids.add(case_id)
        convert_workflow(case_id, raw, is_case=True)

    workflow_map = {
        workflow["workflowId"]: (workflow, own_required)
        for workflow, own_required in workflow_rows
    }

    def transitive_required(workflow_id: str, stack: tuple[str, ...] = ()) -> set[str]:
        if workflow_id in stack:
            raise LegacyMigrationError(
                f"Recursive flow reference: {' -> '.join((*stack, workflow_id))}"
            )
        workflow, own_required = workflow_map[workflow_id]
        required = set(own_required)
        for phase in (
            workflow.get("x-testkit-setup", []),
            workflow.get("steps", []),
            workflow.get("x-testkit-cleanup", []),
        ):
            for step in phase:
                child_id = step.get("workflowId")
                if child_id:
                    if child_id not in workflow_map:
                        raise LegacyMigrationError(
                            f"Unknown referenced flow: {child_id}"
                        )
                    required.update(
                        transitive_required(child_id, (*stack, workflow_id))
                    )
        return required

    all_required: set[str] = set()
    for workflow, _ in workflow_rows:
        required = transitive_required(workflow["workflowId"])
        all_required.update(required)
        properties = {
            **{
                f"project_{name}": {"default": value}
                for name, value in project_vars.items()
            },
            **{name: {"type": "string"} for name in sorted(required)},
        }
        workflow["inputs"] = {
            "type": "object",
            "required": sorted(required),
            "properties": properties,
        }
    try:
        source_url = str(schema_path.relative_to(output_path.parent))
    except ValueError:
        source_url = str(schema_path)
    document = {
        "arazzo": "1.1.0",
        "info": {
            "title": f"{project.get('name', 'legacy')} migrated workflows",
            "version": "1.0.0",
        },
        "sourceDescriptions": [
            {"name": "api", "url": source_url, "type": "openapi"}
        ],
        "workflows": [workflow for workflow, _ in workflow_rows],
    }
    if output_path in {project_path, schema_path}:
        raise LegacyMigrationError(
            "Migration output must not overwrite project.yaml or the OpenAPI schema"
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    serialized = yaml.safe_dump(document, sort_keys=False, allow_unicode=True)
    temp_name: str | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            dir=output_path.parent,
            prefix=f".{output_path.name}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            handle.write(serialized)
            handle.flush()
            os.fsync(handle.fileno())
            temp_name = handle.name
        os.replace(temp_name, output_path)
    except OSError as exc:
        if temp_name:
            Path(temp_name).unlink(missing_ok=True)
        raise LegacyMigrationError(f"Unable to write migration output: {exc}") from exc
    return {
        "schema_version": 1,
        "status": "migrated",
        "source": str(project_path),
        "output": str(output_path),
        "workflow_count": len(workflow_rows),
        "required_environment": sorted(all_required),
        "unsupported_features": [],
    }
