#!/usr/bin/env python3
"""
TestSpec Excel 用例生成脚本：根据 testcases.json 生成 .xlsx 测试用例文档。

用法：
    python generate_excel.py --input testcases.json --output artifacts/cases.xlsx
"""
import argparse
import logging
import os
import sys

try:
    from utils import configure_logging, load_and_validate_testcases
except ImportError:
    from .utils import configure_logging, load_and_validate_testcases

logger = logging.getLogger(__name__)

# 表头定义：(名称, 列宽)
COLUMNS = [
    ("编号", 18),
    ("用例标题", 45),
    ("级别", 8),
    ("预置条件", 25),
    ("操作步骤", 40),
    ("测试预期内容", 35),
    ("执行结果", 12),
    ("执行人", 10),
    ("执行日期", 14),
    ("备注", 20),
]


def _write_sheet(sheet, test_cases: list, header_font, header_fill,
                  Alignment, get_column_letter) -> None:
    """向工作表写入表头和用例数据。"""

    for col_idx, (header, _) in enumerate(COLUMNS, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, (_, width) in enumerate(COLUMNS, 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, tc in enumerate(test_cases, 2):
        tc_id = tc.get("id") or tc.get("case_id") or f"TC-{row_idx - 1:03d}"
        title = tc.get("title") or tc.get("name", "")
        priority = tc.get("priority", "P1")
        preconditions = tc.get("preconditions", "")
        steps = tc.get("steps", "")
        expected = tc.get("expected_result", tc.get("expected", ""))

        sheet.cell(row=row_idx, column=1, value=tc_id)
        sheet.cell(row=row_idx, column=2, value=title)
        sheet.cell(row=row_idx, column=3, value=priority)
        sheet.cell(row=row_idx, column=4, value=preconditions)
        sheet.cell(row=row_idx, column=5, value=steps)
        sheet.cell(row=row_idx, column=6, value=expected)
        for col in range(7, 11):
            sheet.cell(row=row_idx, column=col, value="")


def create_excel_with_openpyxl(test_cases: list, output_path: str) -> None:
    """使用 openpyxl 创建 Excel 文件，包含全量用例和冒烟用例两个 Sheet。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("需要安装 openpyxl，请运行：pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    try:
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        all_sheet = wb.active
        if all_sheet is None:
            all_sheet = wb.create_sheet()
        all_sheet.title = "测试用例"
        _write_sheet(all_sheet, test_cases, header_font, header_fill, Alignment, get_column_letter)

        smoke_cases = [tc for tc in test_cases if tc.get("type") == "冒烟"]
        if smoke_cases:
            smoke_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
            smoke_font = Font(bold=True, color="FFFFFF", size=11)
            smoke_sheet = wb.create_sheet(title="冒烟用例")
            _write_sheet(smoke_sheet, smoke_cases, smoke_font, smoke_fill, Alignment, get_column_letter)

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
    finally:
        wb.close()


def main() -> None:
    configure_logging()

    parser = argparse.ArgumentParser(description="Generate Excel test cases from JSON")
    parser.add_argument("--input", "-i", required=True, help="Path to testcases.json")
    parser.add_argument("--output", "-o", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    test_cases = load_and_validate_testcases(args.input)
    create_excel_with_openpyxl(test_cases, args.output)
    print(f"已生成：{args.output}")


if __name__ == "__main__":
    main()
