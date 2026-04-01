#!/usr/bin/env python3
"""Export YAML/JSON API specs to Excel."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from _spec_table import COLUMNS, load_spec, spec_to_rows

from openpyxl import Workbook
from openpyxl.styles import Font


def main():
    parser = argparse.ArgumentParser(description="Export YAML/JSON API specs to Excel")
    parser.add_argument("--input", "-i", required=True, help="Input YAML/JSON spec path")
    parser.add_argument("--output", "-o", required=True, help="Output Excel path")
    args = parser.parse_args()

    inp = Path(args.input)
    if not inp.exists():
        raise FileNotFoundError(f"Input file not found: {inp}")

    rows = spec_to_rows(load_spec(inp))

    wb = Workbook()
    ws = wb.active
    ws.title = "用例"
    for ci, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = Font(bold=True)
    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(COLUMNS, 1):
            ws.cell(row=ri, column=ci, value=row.get(col, ""))

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Exported {len(rows)} rows to {out}")


if __name__ == "__main__":
    main()
