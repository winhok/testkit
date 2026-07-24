#!/usr/bin/env python3
"""Import legacy cases into an isolated, unverified TestSpec staging artifact."""
from __future__ import annotations

import argparse
import csv
import json
import re
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter
from pathlib import Path
from typing import Any, Iterable


FIELD_ALIASES = {
    "id": ("id", "case_id", "编号"),
    "title": ("title", "name", "标题", "用例标题", "用例名称", "场景名称"),
    "priority": ("priority", "优先级", "级别"),
    "preconditions": ("preconditions", "前置条件", "预置条件"),
    "steps": ("steps", "步骤", "操作步骤", "测试步骤"),
    "expected_result": (
        "expected_result",
        "expected",
        "预期",
        "期望结果",
        "预期结果",
        "测试预期内容",
    ),
    "type": ("type", "类型"),
    "feature": ("feature", "功能", "模块"),
}
ALIASES = {
    alias.strip().lower(): canonical
    for canonical, aliases in FIELD_ALIASES.items()
    for alias in aliases
}
SAFE_SOURCE_LABEL = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,63}$")


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def canonical_row(row: dict[str, Any]) -> dict[str, str]:
    result = {field: "" for field in FIELD_ALIASES}
    for key, value in row.items():
        canonical = ALIASES.get(clean(key).lower())
        if canonical and not result[canonical]:
            result[canonical] = clean(value)
    return result


def split_markdown_row(line: str) -> list[str]:
    stripped = line.strip()
    if stripped.startswith("|"):
        stripped = stripped[1:]
    if stripped.endswith("|") and not stripped.endswith(r"\|"):
        stripped = stripped[:-1]
    return [
        part.replace(r"\|", "|").strip()
        for part in re.split(r"(?<!\\)\|", stripped)
    ]


def markdown_rows(path: Path) -> list[tuple[int, dict[str, Any]]]:
    lines = path.read_text(encoding="utf-8").splitlines()
    rows: list[tuple[int, dict[str, Any]]] = []
    index = 0
    while index < len(lines) - 1:
        if "|" not in lines[index]:
            index += 1
            continue
        headers = split_markdown_row(lines[index])
        separator = split_markdown_row(lines[index + 1])
        if (
            len(headers) != len(separator)
            or not separator
            or not all(re.fullmatch(r":?-{3,}:?", item.replace(" ", "")) for item in separator)
        ):
            index += 1
            continue
        index += 2
        while index < len(lines) and "|" in lines[index] and lines[index].strip():
            values = split_markdown_row(lines[index])
            rows.append((index + 1, dict(zip(headers, values))))
            index += 1
    if rows:
        return rows

    blocks = re.split(r"(?m)^#{2,4}\s+", path.read_text(encoding="utf-8"))
    for block_index, block in enumerate(blocks[1:], start=1):
        block_lines = [line.rstrip() for line in block.strip().splitlines()]
        if not block_lines:
            continue
        title = block_lines[0].strip()
        row = parse_labeled_lines(block_lines[1:])
        row["title"] = row.get("title") or title
        rows.append((block_index, row))
    return rows


def parse_labeled_lines(lines: list[str]) -> dict[str, str]:
    row: dict[str, str] = {}
    current = ""
    values: dict[str, list[str]] = {}
    for raw_line in lines:
        line = raw_line.strip().lstrip("-* ").strip()
        if not line:
            continue
        match = re.match(r"^([^:：]{1,24})[:：]\s*(.*)$", line)
        if match:
            canonical = ALIASES.get(match.group(1).strip().lower())
            if canonical:
                current = canonical
                values.setdefault(current, [])
                if match.group(2).strip():
                    values[current].append(match.group(2).strip())
                continue
            current = ""
            continue
        if current:
            values[current].append(line)
    for key, parts in values.items():
        row[key] = "\n".join(parts).strip()
    return row


def text_rows(path: Path) -> list[tuple[int, dict[str, Any]]]:
    blocks = re.split(r"\n\s*\n+", path.read_text(encoding="utf-8").strip())
    rows: list[tuple[int, dict[str, Any]]] = []
    for index, block in enumerate(blocks, start=1):
        lines = [line.rstrip() for line in block.splitlines() if line.strip()]
        if not lines:
            continue
        row = parse_labeled_lines(lines)
        first_is_label = bool(re.match(r"^([^:：]{1,24})[:：]", lines[0].strip()))
        if not row.get("title") and not first_is_label:
            row["title"] = lines[0].strip().lstrip("#").strip()
        rows.append((index, row))
    return rows


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def xml_topic(node: ET.Element) -> dict[str, Any]:
    title = ""
    notes = ""
    children: list[dict[str, Any]] = []
    for child in node:
        name = local_name(child.tag)
        if name == "title" and child.text:
            title = child.text.strip()
        elif name == "notes":
            for value in child.iter():
                if local_name(value.tag) == "plain" and value.text:
                    notes = value.text.strip()
                    break
        elif name == "children":
            for descendant in child.iter():
                if local_name(descendant.tag) == "topics":
                    children = [
                        xml_topic(topic)
                        for topic in list(descendant)
                        if local_name(topic.tag) == "topic"
                    ]
                    break
    return {"title": title, "notes": notes, "children": children}


def json_topic(node: Any) -> dict[str, Any]:
    if not isinstance(node, dict):
        return {"title": "", "notes": "", "children": []}
    title = clean(node.get("title"))
    notes_raw = node.get("notes")
    if isinstance(notes_raw, dict):
        notes = clean(notes_raw.get("plain") or notes_raw.get("content"))
    else:
        notes = clean(notes_raw)
    children_raw = node.get("children") or node.get("topics") or []
    if isinstance(children_raw, dict):
        children_raw = (
            children_raw.get("attached")
            or children_raw.get("topics")
            or []
        )
    if not isinstance(children_raw, list):
        children_raw = []
    return {
        "title": title,
        "notes": notes,
        "children": [json_topic(child) for child in children_raw],
    }


def topic_value(node: dict[str, Any]) -> str:
    if clean(node.get("notes")):
        return clean(node["notes"])
    leaves: list[str] = []

    def walk(item: dict[str, Any]) -> None:
        children = item.get("children") or []
        if not children:
            if clean(item.get("title")):
                leaves.append(clean(item["title"]))
            return
        for child in children:
            walk(child)

    walk(node)
    return "\n".join(leaves)


def rows_from_topic(root: dict[str, Any]) -> list[tuple[int, dict[str, Any]]]:
    rows: list[tuple[int, dict[str, Any]]] = []
    case_fields = {"title", "preconditions", "steps", "expected_result"}

    def walk(node: dict[str, Any]) -> None:
        children = node.get("children") or []
        labeled: dict[str, str] = {}
        for child in children:
            canonical = ALIASES.get(clean(child.get("title")).lower())
            if canonical and canonical not in labeled:
                labeled[canonical] = topic_value(child)
        if case_fields.intersection(labeled) and clean(node.get("title")):
            labeled.setdefault("title", clean(node["title"]))
            rows.append((len(rows) + 1, labeled))
            return
        for child in children:
            walk(child)

    walk(root)
    return rows


def xmind_rows(path: Path) -> list[tuple[int, dict[str, Any]]]:
    with zipfile.ZipFile(path) as archive:
        names = set(archive.namelist())
        roots: list[dict[str, Any]] = []
        if "content.json" in names:
            raw = json.loads(archive.read("content.json").decode("utf-8"))
            sheets = raw if isinstance(raw, list) else [raw]
            for sheet in sheets:
                if not isinstance(sheet, dict):
                    continue
                roots.append(json_topic(sheet.get("rootTopic") or sheet))
        elif "content.xml" in names:
            document = ET.fromstring(archive.read("content.xml"))
            for sheet in document.iter():
                if local_name(sheet.tag) != "sheet":
                    continue
                root_topic = next(
                    (
                        child
                        for child in sheet
                        if local_name(child.tag) == "topic"
                    ),
                    None,
                )
                if root_topic is not None:
                    roots.append(xml_topic(root_topic))
        else:
            raise ValueError("XMind input must contain content.json or content.xml")
    rows: list[tuple[int, dict[str, Any]]] = []
    for root in roots:
        for _, row in rows_from_topic(root):
            rows.append((len(rows) + 1, row))
    return rows


def load_rows(path: Path) -> list[tuple[int, dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [(index, dict(row)) for index, row in enumerate(csv.DictReader(handle), start=2)]
    if suffix == ".json":
        raw = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(raw, dict):
            raw = raw.get("testcases", raw.get("cases"))
        if not isinstance(raw, list):
            raise ValueError("JSON input must be an array or contain a testcases/cases array")
        return [(index, row) for index, row in enumerate(raw, start=1) if isinstance(row, dict)]
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency guidance
            raise RuntimeError("openpyxl is required to import .xlsx files") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [clean(value) for value in next(rows, ())]
        return [
            (index, dict(zip(headers, values)))
            for index, values in enumerate(rows, start=2)
        ]
    if suffix == ".md":
        return markdown_rows(path)
    if suffix == ".txt":
        return text_rows(path)
    if suffix == ".xmind":
        return xmind_rows(path)
    raise ValueError(
        "unsupported input format; expected .xlsx, .csv, .json, .md, .txt, or .xmind"
    )


def normalized_text(value: str) -> str:
    return re.sub(r"[\W_]+", "", value, flags=re.UNICODE).lower()


def duplicate_values(values: Iterable[str]) -> set[str]:
    normalized = [normalized_text(value) for value in values if value]
    counts = Counter(normalized)
    return {value for value, count in counts.items() if count > 1}


def import_cases(input_path: Path, source_label: str) -> dict[str, Any]:
    if not SAFE_SOURCE_LABEL.fullmatch(source_label):
        raise ValueError(
            "source label must be a non-sensitive label using letters, digits, dot, underscore, or hyphen"
        )
    rows = load_rows(input_path)
    imported: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = (
        [{"type": "no_parseable_rows", "source_row": 0}]
        if not rows
        else []
    )
    used_case_ids: set[str] = set()
    skipped = 0

    for sequence, (source_row, raw) in enumerate(rows, start=1):
        row = canonical_row(raw)
        if not any(row.values()):
            skipped += 1
            continue
        missing = [
            field
            for field in ("title", "steps", "expected_result")
            if not row[field]
        ]
        if missing:
            warnings.append({
                "type": "missing_key_fields",
                "source_row": source_row,
                "fields": missing,
            })
        source_case_id = row["id"]
        case_id = source_case_id or f"LEGACY_{sequence:04d}"
        if case_id in used_case_ids:
            base_id = case_id
            duplicate_number = 2
            while case_id in used_case_ids:
                case_id = f"{base_id}__DUP_{duplicate_number}"
                duplicate_number += 1
            warnings.append({
                "type": "duplicate_source_id_candidate",
                "case_id": case_id,
                "source_case_id": source_case_id,
                "source_row": source_row,
            })
        used_case_ids.add(case_id)
        origin: dict[str, Any] = {
            "kind": "legacy-import",
            "source_label": source_label,
            "source_row": source_row,
            "source_format": input_path.suffix.lower().lstrip("."),
        }
        if source_case_id:
            origin["source_case_id"] = source_case_id
        imported.append({
            "id": case_id,
            "title": row["title"] or f"待补充标题_{sequence:04d}",
            "priority": row["priority"] or "P2",
            "type": row["type"] or "其他",
            "feature": row["feature"] or "待分类",
            "preconditions": row["preconditions"],
            "steps": row["steps"],
            "expected_result": row["expected_result"],
            "tp_refs": [],
            "origin": origin,
            "trust": {
                "status": "unverified",
                "reason": "pending-current-prd-reconciliation",
            },
            "reconciliation": {
                "status": "unresolved",
                "requirement_refs": [],
            },
        })

    duplicate_titles = duplicate_values(case["title"] for case in imported)
    duplicate_bodies = duplicate_values(
        f"{case['preconditions']}|{case['steps']}|{case['expected_result']}"
        for case in imported
    )
    for case in imported:
        if normalized_text(case["title"]) in duplicate_titles:
            warnings.append({
                "type": "duplicate_title_candidate",
                "case_id": case["id"],
                "source_row": case["origin"]["source_row"],
            })
        body = normalized_text(
            f"{case['preconditions']}|{case['steps']}|{case['expected_result']}"
        )
        if body and body in duplicate_bodies:
            warnings.append({
                "type": "duplicate_body_candidate",
                "case_id": case["id"],
                "source_row": case["origin"]["source_row"],
            })

    return {
        "schema_version": 2,
        "_context": {
            "source_skill": "testspec-import",
            "canonical_source_policy": "prd-first",
            "publish_eligibility": "blocked",
            "origin": {"kind": "legacy-import"},
            "trust": {"status": "unverified"},
        },
        "import_summary": {
            "input_rows": len(rows),
            "imported_cases": len(imported),
            "skipped_rows": skipped,
            "warning_count": len(warnings),
        },
        "warnings": warnings,
        "testcases": imported,
    }


def build_reconciliation(imported: dict[str, Any]) -> dict[str, Any]:
    records = [
        {
            "legacy_case_id": case["id"],
            "source_row": case["origin"]["source_row"],
            "status": "unresolved",
            "requirement_refs": [],
            "question_refs": [],
            "replacement_candidate_id": "",
            "reason": "",
        }
        for case in imported["testcases"]
    ]
    return {
        "schema_version": 1,
        "_context": {
            "source_skill": "testspec-import",
            "canonical_source_policy": "prd-first",
            "status": "pending",
        },
        "summary": {
            "keep": 0,
            "revise": 0,
            "merge": 0,
            "retire": 0,
            "unresolved": len(records),
        },
        "records": records,
    }


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        required=True,
        help="Legacy .xlsx, .csv, .json, .md, .txt, or .xmind input",
    )
    parser.add_argument("--output", required=True, help="Isolated staging JSON output")
    parser.add_argument(
        "--reconciliation-output",
        help="Reconciliation JSON output; defaults to reconciliation.json beside --output",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Explicitly replace existing staging outputs",
    )
    parser.add_argument(
        "--source-label",
        default="legacy-source",
        help="Non-sensitive logical label stored in provenance; never use a local path",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    reconciliation_path = (
        Path(args.reconciliation_output)
        if args.reconciliation_output
        else output_path.with_name("reconciliation.json")
    )
    if output_path == reconciliation_path:
        parser.error("--output and --reconciliation-output must be different paths")
    existing = [path.name for path in (output_path, reconciliation_path) if path.exists()]
    if existing and not args.overwrite:
        parser.error(
            "refusing to overwrite existing staging output(s): "
            + ", ".join(existing)
            + "; pass --overwrite only after explicit authorization"
        )

    result = import_cases(input_path, args.source_label)
    reconciliation = build_reconciliation(result)
    write_json(output_path, result)
    write_json(reconciliation_path, reconciliation)
    print(json.dumps({
        **result["import_summary"],
        "reconciliation_records": len(reconciliation["records"]),
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
