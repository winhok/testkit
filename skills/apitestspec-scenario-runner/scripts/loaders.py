"""Self-contained spec file loaders and normalizers.

Handles YAML, JSON, CSV, and Excel case specs, project discovery, and flow loading.
"""
from __future__ import annotations

import csv as _csv
import json
import sys
from pathlib import Path
from typing import Any

import yaml

sys.path.insert(0, str(Path(__file__).resolve().parent))

from engine import CaseDocument, CaseSpec, FlowSpec, ProjectSpec, StepSpec


class SpecValidationError(ValueError):
    pass

# ---------------------------------------------------------------------------
# Normalizer
# ---------------------------------------------------------------------------

def _expect_obj(v: Any, *, loc: str) -> dict:
    if not isinstance(v, dict):
        raise SpecValidationError(f"{loc} must be an object")
    return v


def _expect_list(v: Any, *, loc: str) -> list:
    if not isinstance(v, list):
        raise SpecValidationError(f"{loc} must be a list")
    return v


def _norm_step(raw: dict, *, loc: str) -> StepSpec:
    s = _expect_obj(raw, loc=loc)
    has_req = s.get("request") is not None
    has_use = s.get("use") is not None
    has_sleep = s.get("sleep") is not None
    if not has_req and not has_use and not has_sleep:
        raise SpecValidationError(f"{loc} must define request, use, or sleep")
    if has_req and has_use:
        raise SpecValidationError(f"{loc} must not define both request and use")
    return StepSpec(
        name=s.get("name"), use=s.get("use"), inputs=s.get("inputs"),
        save_as=s.get("save_as"), request=s.get("request"),
        extract=s.get("extract"), validate=s.get("validate"), sleep=s.get("sleep"),
    )


def _norm_steps(raw: Any, *, loc: str) -> list[StepSpec]:
    return [_norm_step(s, loc=f"{loc}[{i}]") for i, s in enumerate(_expect_list(raw or [], loc=loc))]


def _norm_flow(name: str, val: Any) -> FlowSpec:
    obj = _expect_obj(val, loc=f"flows.{name}")
    return FlowSpec(name=name, steps=_norm_steps(obj.get("steps", []), loc=f"flows.{name}.steps"))


def _norm_case(val: Any, *, idx: int) -> CaseSpec:
    loc = f"cases[{idx}]"
    c = _expect_obj(val, loc=loc)
    if "id" not in c:
        raise SpecValidationError(f"{loc}.id is required")
    if "name" not in c:
        raise SpecValidationError(f"{loc}.name is required")
    on_f = c.get("on_failure", "stop")
    if on_f not in {"stop", "continue"}:
        raise SpecValidationError(f"{loc}.on_failure must be stop or continue")
    return CaseSpec(
        id=c["id"], name=c["name"], tags=list(c.get("tags") or []),
        continue_on_failure=bool(c.get("continue_on_failure", False)), on_failure=on_f,
        setup=_norm_steps(c.get("setup", []), loc=f"{loc}.setup"),
        steps=_norm_steps(c.get("steps", []), loc=f"{loc}.steps"),
        teardown=_norm_steps(c.get("teardown", []), loc=f"{loc}.teardown"),
    )


def normalize_document(raw: dict[str, Any], source: str | None = None) -> CaseDocument:
    doc = _expect_obj(raw, loc="document")
    raw_flows = doc.get("flows", {})
    if not isinstance(raw_flows, dict):
        raise SpecValidationError("flows must be an object")
    flows = {k: _norm_flow(k, v) for k, v in raw_flows.items()}
    raw_cases = doc.get("cases", [])
    if not isinstance(raw_cases, list):
        raise SpecValidationError("cases must be a list")
    cases = [_norm_case(c, idx=i) for i, c in enumerate(raw_cases)]
    return CaseDocument(cases=cases, flows=flows, source=source)

# ---------------------------------------------------------------------------
# File readers
# ---------------------------------------------------------------------------

def _read_mapping(path: Path) -> dict[str, Any]:
    if path.suffix.lower() == ".json":
        return json.loads(path.read_text(encoding="utf-8"))
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


_CASE_SUFFIXES = {".yaml", ".yml", ".json", ".csv", ".xlsx", ".xls"}
_MAP_SUFFIXES = {".yaml", ".yml", ".json"}


import re as _re

_EXPR_RE = _re.compile(
    r"(\$[\w.]+)\s*(==|!=|>=|<=|>|<|\s+contains\s+)\s*(.+)"
)


def _parse_check_expr(raw: str) -> list[dict]:
    """Parse compound expression like '$.code==200 && $.data.token' into validate list."""
    parts = [p.strip() for p in raw.split("&&")]
    result: list[dict] = []
    op_map = {"==": "eq", "!=": "ne", ">": "gt", ">=": "gte", "<": "lt", "<=": "lte"}
    for part in parts:
        if not part:
            continue
        m = _EXPR_RE.match(part)
        if m:
            field, op, value = m.group(1), m.group(2).strip(), m.group(3).strip()
            if op == "contains":
                result.append({"contains": [field, value]})
            else:
                mapped = op_map.get(op, "eq")
                try:
                    typed_val: object = int(value)
                except ValueError:
                    try:
                        typed_val = float(value)
                    except ValueError:
                        typed_val = value
                result.append({mapped: [field, typed_val]})
        else:
            result.append({"exists": part})
    return result


def _parse_extract_expr(raw: str) -> dict[str, str]:
    """Parse 'token=$.data.token ; user_id=$.data.id' into extract dict."""
    result: dict[str, str] = {}
    for item in raw.split(";"):
        item = item.strip()
        if "=" not in item:
            continue
        key, val = item.split("=", 1)
        result[key.strip()] = val.strip()
    return result


def _csv_row_to_step(row: dict[str, str]) -> dict:
    """Convert a single CSV/Excel row into a step dict (request step or use step)."""
    use = (row.get("前置依赖") or "").strip()
    url = (row.get("接口路径") or "").strip()

    step_name_raw = (row.get("用例名称") or "step")
    step_name = step_name_raw.split(" / ", 1)[1] if " / " in step_name_raw else step_name_raw

    if use and not url:
        step: dict[str, object] = {"name": step_name, "use": use}
        return step

    req: dict[str, object] = {"method": (row.get("请求方法") or "GET").strip(), "url": url}
    body = (row.get("请求体/参数") or "").strip()
    if body:
        try:
            req["json"] = json.loads(body)
        except json.JSONDecodeError:
            req["raw_body"] = body

    validations: list[dict] = []
    status_str = (row.get("预期状态码") or "").strip()
    if status_str:
        validations.append({"eq": ["status_code", int(status_str)]})
    check = (row.get("预期响应校验") or "").strip()
    if check:
        validations.extend(_parse_check_expr(check))

    extract_raw = (row.get("依赖产出提取") or "").strip()
    extract = _parse_extract_expr(extract_raw) if extract_raw else None

    step = {"name": step_name, "request": req, "validate": validations}
    if use:
        step = {"name": step_name, "use": use}
        return step
    if extract:
        step["extract"] = extract
    return step


def _aggregate_rows_to_cases(rows: list[dict[str, str]]) -> list[dict]:
    """Group rows by 用例ID into multi-step cases with setup/steps/teardown."""
    from collections import OrderedDict
    groups: OrderedDict[str, list[dict[str, str]]] = OrderedDict()
    for row in rows:
        cid = (row.get("用例ID") or "").strip()
        if not cid:
            continue
        groups.setdefault(cid, []).append(row)

    cases: list[dict] = []
    for cid, group_rows in groups.items():
        first = group_rows[0]
        case_name_raw = (first.get("用例名称") or cid)
        case_name = case_name_raw.split(" / ", 1)[0] if " / " in case_name_raw else case_name_raw
        setup: list[dict] = []
        steps: list[dict] = []
        teardown: list[dict] = []
        for r in group_rows:
            phase = (r.get("注入方式") or "").strip().lower()
            step = _csv_row_to_step(r)
            if phase == "setup":
                setup.append(step)
            elif phase == "teardown":
                teardown.append(step)
            else:
                steps.append(step)
        case: dict[str, object] = {"id": cid, "name": case_name, "steps": steps}
        if setup:
            case["setup"] = setup
        if teardown:
            case["teardown"] = teardown
        cases.append(case)
    return cases


def _load_csv(path: Path) -> CaseDocument:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(_csv.DictReader(f))
    cases = _aggregate_rows_to_cases(rows)
    return normalize_document({"cases": cases, "flows": {}}, source=str(path))


def _load_excel(path: Path) -> CaseDocument:
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    raw_rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c else "" for c in raw_rows[0]]
    dict_rows = []
    for row in raw_rows[1:]:
        rec = {headers[i]: ("" if row[i] is None else str(row[i])) for i in range(len(headers)) if headers[i]}
        dict_rows.append(rec)
    wb.close()
    cases = _aggregate_rows_to_cases(dict_rows)
    return normalize_document({"cases": cases, "flows": {}}, source=str(path))


def load_case_file(path: Path) -> CaseDocument:
    sfx = path.suffix.lower()
    if sfx in {".yaml", ".yml", ".json"}:
        return normalize_document(_read_mapping(path), source=str(path))
    if sfx == ".csv":
        return _load_csv(path)
    if sfx in {".xlsx", ".xls"}:
        return _load_excel(path)
    raise ValueError(f"Unsupported file type: {path.suffix}")

# ---------------------------------------------------------------------------
# Project loader
# ---------------------------------------------------------------------------

def load_project_spec(path: Path) -> ProjectSpec:
    raw = _read_mapping(path)
    p = raw.get("project", raw)
    if "name" not in p:
        raise SpecValidationError("project.name is required")
    if "base_url" not in p:
        raise SpecValidationError("project.base_url is required")
    return ProjectSpec(
        name=p["name"], base_url=p["base_url"],
        vars=dict(p.get("vars") or {}), defaults=dict(p.get("defaults") or {}),
        report=dict(p.get("report") or {}),
    )


def _configured_dir(project_path: Path, project: ProjectSpec, key: str, default: str) -> Path:
    root = project_path.parent
    cfg = project.report.get(key) if isinstance(project.report, dict) else None
    return root / str(cfg) if cfg else root / default


def _iter_files(root: Path, suffixes: set[str]) -> list[Path]:
    if not root.exists() or not root.is_dir():
        return []
    return [p for p in sorted(root.rglob("*")) if p.suffix.lower() in suffixes]


def discover_flows(project_path: Path, project: ProjectSpec) -> dict[str, FlowSpec]:
    cfg_root = _configured_dir(project_path, project, "flows_dir", "flows")
    roots = [cfg_root]
    default_root = project_path.parent / "flows"
    if cfg_root != default_root:
        roots.append(default_root)
    discovered: dict[str, FlowSpec] = {}
    seen: set[Path] = set()
    for root in roots:
        for p in _iter_files(root, _MAP_SUFFIXES):
            if p in seen:
                continue
            seen.add(p)
            doc = normalize_document(_read_mapping(p), source=str(p))
            discovered.update(doc.flows)
    return discovered


def discover_case_files(project_path: Path, project: ProjectSpec) -> list[Path]:
    return _iter_files(_configured_dir(project_path, project, "cases_dir", "cases"), _CASE_SUFFIXES)
